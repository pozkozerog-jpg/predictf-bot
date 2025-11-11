"""
Модуль для автоматического обновления Excel файла со статистикой пользователей
"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from modules.database import get_all_users_for_export
import os


EXCEL_FILENAME = "users_stats.xlsx"


def update_excel_file():
    """
    Автоматически обновляет Excel файл со списком всех пользователей бота
    Создает файл, если его нет, или обновляет существующий
    """
    try:
        # Получаем всех пользователей из базы данных
        users = get_all_users_for_export()
        
        # Создаем новую рабочую книгу
        wb = Workbook()
        ws = wb.active
        ws.title = "Пользователи бота"
        
        # Заголовок
        ws['A1'] = "⚽ ПОЛЬЗОВАТЕЛИ FOOTBALL PREDICTOR BOT"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1F77B4", end_color="1F77B4", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells('A1:H1')
        ws.row_dimensions[1].height = 30
        
        # Дата обновления
        ws['A2'] = f"Обновлено: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}"
        ws['A2'].font = Font(italic=True, size=10)
        ws['A2'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A2:H2')
        
        # Заголовки столбцов
        headers = [
            "№",
            "User ID",
            "Username",
            "Имя",
            "Фамилия",
            "Первое использование",
            "Последнее использование",
            "Всего использований"
        ]
        
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Пишем заголовки
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_style
        
        # Заполняем данные пользователей
        row = 5
        for idx, user in enumerate(users, start=1):
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=user.get('user_id'))
            ws.cell(row=row, column=3, value=user.get('username') or '-')
            ws.cell(row=row, column=4, value=user.get('first_name') or '-')
            ws.cell(row=row, column=5, value=user.get('last_name') or '-')
            
            # Форматируем даты
            first_seen = user.get('first_seen')
            last_seen = user.get('last_seen')
            if first_seen:
                ws.cell(row=row, column=6, value=first_seen.strftime("%d.%m.%Y %H:%M"))
            if last_seen:
                ws.cell(row=row, column=7, value=last_seen.strftime("%d.%m.%Y %H:%M"))
            
            ws.cell(row=row, column=8, value=user.get('total_actions'))
            
            # Применяем границы и выравнивание
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = border_style
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
            
            row += 1
        
        # Устанавливаем ширину столбцов
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 15
        
        # Добавляем итоговую строку
        total_row = row + 1
        ws.cell(row=total_row, column=1, value="ИТОГО:")
        ws.cell(row=total_row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=total_row, column=8, value=len(users))
        ws.cell(row=total_row, column=8).font = Font(bold=True, size=12)
        ws.merge_cells(f'A{total_row}:G{total_row}')
        ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center")
        
        # Сохраняем файл
        wb.save(EXCEL_FILENAME)
        print(f"✅ Excel файл обновлен: {EXCEL_FILENAME} ({len(users)} пользователей)")
        
    except Exception as e:
        print(f"❌ Ошибка при обновлении Excel файла: {e}")
        import traceback
        traceback.print_exc()
